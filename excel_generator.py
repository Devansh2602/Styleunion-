"""
Generate Invoice + Packing List Excel by filling the real template.
All borders and formatting come from template.xlsx (converted from the .xls template).
"""

import os
from copy import copy as dcopy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

SIZE_ORDER = {'XS': 0, 'S': 1, 'M': 2, 'L': 3, 'XL': 4, 'XXL': 5, '2XL': 5, '3XL': 6}

TEMPLATE_INV_ITEM_ROWS = 5    # rows 23-27 in template
TEMPLATE_PKG_CARTON_ROWS = 37 # rows 17-53 in template


# ── helpers ────────────────────────────────────────────────────────────────

def _sort_items(items):
    return sorted(items, key=lambda x: SIZE_ORDER.get(x['size'], 99))


def _num_to_words_indian(n):
    n = int(round(n))
    if n == 0:
        return 'ZERO'
    ones = [
        '', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE',
        'TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN',
        'SEVENTEEN', 'EIGHTEEN', 'NINETEEN',
    ]
    tens = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY',
            'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']

    def two_dig(x):
        return ones[x] if x < 20 else tens[x // 10] + (' ' + ones[x % 10] if x % 10 else '')

    def three_dig(x):
        if x >= 100:
            return ones[x // 100] + ' HUNDRED' + (' ' + two_dig(x % 100) if x % 100 else '')
        return two_dig(x)

    parts = []
    if n >= 10_000_000:
        parts.append(three_dig(n // 10_000_000) + ' CRORE')
        n %= 10_000_000
    if n >= 100_000:
        parts.append(three_dig(n // 100_000) + ' LAKH')
        n %= 100_000
    if n >= 1_000:
        parts.append(two_dig(n // 1_000) + ' THOUSAND')
        n %= 1_000
    if n > 0:
        parts.append(three_dig(n))
    return ' '.join(parts) + ' ONLY'


def _calculate_cartons(items, capacity):
    cartons = []
    carton_no = 1
    for item in _sort_items(items):
        qty = item['qty']
        full, rem = divmod(qty, capacity)
        article_size = f"{item['barcode']} - {item['size']}"
        for _ in range(full):
            cartons.append({'no': carton_no, 'article_size': article_size,
                            'description': item['description'], 'color': item['color'],
                            'qty': capacity})
            carton_no += 1
        if rem:
            cartons.append({'no': carton_no, 'article_size': article_size,
                            'description': item['description'], 'color': item['color'],
                            'qty': rem})
            carton_no += 1
    return cartons


def _save_row_styles(ws, row_num, num_cols):
    """Return list of copied style dicts for each cell in the row."""
    return [
        {
            'font': dcopy(ws.cell(row_num, c).font),
            'border': dcopy(ws.cell(row_num, c).border),
            'fill': dcopy(ws.cell(row_num, c).fill),
            'alignment': dcopy(ws.cell(row_num, c).alignment),
            'number_format': ws.cell(row_num, c).number_format,
        }
        for c in range(1, num_cols + 1)
    ]


def _apply_row_styles(ws, row_num, styles):
    for col, s in enumerate(styles, 1):
        c = ws.cell(row_num, col)
        c.font = dcopy(s['font'])
        c.border = dcopy(s['border'])
        c.fill = dcopy(s['fill'])
        c.alignment = dcopy(s['alignment'])
        c.number_format = s['number_format']


def _w(ws, row, col, value):
    ws.cell(row, col).value = value


def _clear_col_borders(ws, col):
    """Remove all borders from every cell in a column (clears stray template borders)."""
    empty = Border()
    for row in range(1, ws.max_row + 1):
        ws.cell(row, col).border = empty


# ── INV sheet filler ───────────────────────────────────────────────────────

def _fill_inv(ws, po, inv_no, inv_date, total_cartons):
    items = _sort_items(po['items'])
    n = len(items)
    total_qty = sum(i['qty'] for i in items)
    total_amount = sum(i['amount'] for i in items)
    igst_rate = po.get('igst_rate', 0.05)
    igst_amount = round(total_amount * igst_rate)
    grand_total = total_amount + igst_amount

    consignee = po.get('consignee_name', 'Nexon Omniverse Limited')
    addr = po.get('consignee_address_lines', [
        'AWL INDIA PVT LTD, SV NO 705/2, 706/4 DEVARYAMJAL, Opp.',
        'Sai Geetha Ashramam, Tumkunta Municipality, Medchal Malkajgiri',
        'Hyderabad, 500078',
        'Telangana,INDIA',
    ])
    cgst = po.get('consignee_gst', '36AAQCS3643G1Z2')

    # ── header cells ──────────────────────────────────────────────
    _w(ws, 2, 7,  f'Invoice No : {inv_no}')
    _w(ws, 3, 7,  f'Date of Issue of Invoice :  {inv_date}')
    _w(ws, 9, 7,  f'Purchase Order No : {po["po_number"]}')
    _w(ws, 10, 8, po.get('po_date', ''))
    _w(ws, 12, 7, f'  No. of Cartons : {total_cartons}')

    # ── address block (rows 15–20) ─────────────────────────────────
    top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    _w(ws, 15, 1, consignee + '\n')
    ws.cell(15, 1).alignment = dcopy(top_align)
    _w(ws, 15, 7, consignee + '\n')
    ws.cell(15, 7).alignment = dcopy(top_align)
    for i, line in enumerate(addr[:4], 16):
        _w(ws, i, 1, line)
        _w(ws, i, 7, line)
    _w(ws, 20, 1, f'GST NO :- {cgst}')
    _w(ws, 20, 7, f'GST NO :- {cgst}')

    # ── item rows ─────────────────────────────────────────────────
    FIRST = 23
    TMPL = TEMPLATE_INV_ITEM_ROWS

    # Snapshot style of template item row before it's deleted
    item_styles = _save_row_styles(ws, FIRST, 14)

    ws.delete_rows(FIRST, TMPL)   # TOTAL row shifts to 23
    ws.insert_rows(FIRST, n)      # TOTAL row is now at 23+n

    for idx, item in enumerate(items):
        row = FIRST + idx
        _apply_row_styles(ws, row, item_styles)
        ws.row_dimensions[row].height = 21.75
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        _w(ws, row,  1, idx + 1)
        _w(ws, row,  2, item['barcode'])
        _w(ws, row,  4, item['color'])
        _w(ws, row,  5, item['size'])
        _w(ws, row,  6, item['description'])
        _w(ws, row,  7, item['style'])
        _w(ws, row,  8, item['hsn'])
        _w(ws, row,  9, 'EA')
        _w(ws, row, 10, item['mrp'])
        _w(ws, row, 11, item['rate'])
        _w(ws, row, 12, item['qty'])
        _w(ws, row, 13, item['qty'])
        _w(ws, row, 14, item['amount'])

    # ── totals and tax rows ───────────────────────────────────────
    T = FIRST + n  # TOTAL row (dynamic position)

    _w(ws, T,      12, total_qty)
    _w(ws, T,      13, total_qty)
    _w(ws, T,      14, total_amount)
    _w(ws, T + 1,  14, 0)                # Discount
    _w(ws, T + 2,  14, total_amount)     # Sub-total
    _w(ws, T + 3,  14, 0)                # MRP VALUE
    _w(ws, T + 4,  14, 0)                # Assessable value
    _w(ws, T + 6,  12, igst_rate)        # IGST rate
    _w(ws, T + 6,  14, igst_amount)      # IGST amount
    _w(ws, T + 9,  14, 0)                # Round off
    _w(ws, T + 10, 14, grand_total)      # Grand total
    _w(ws, T + 11, 1,
       f'Amount Chargeable ( in Words ) : {_num_to_words_indian(grand_total)}.')
    _w(ws, T + 12, 1,
       f'IGST Amount ( in words) : {_num_to_words_indian(igst_amount)}.')


# ── PKG sheet filler ───────────────────────────────────────────────────────

def _fill_pkg(ws, po, inv_no, inv_date, cartons, total_cartons):
    total_pcs = sum(c['qty'] for c in cartons)

    consignee = po.get('consignee_name', 'Nexon Omniverse Limited')
    addr = po.get('consignee_address_lines', [
        'AWL INDIA PVT LTD, SV NO 705/2, 706/4 DEVARYAMJAL, Opp.',
        'Sai Geetha Ashramam, Tumkunta Municipality, Medchal Malkajgiri',
        'Hyderabad, 500078',
        'Telangana,INDIA',
    ])
    cgst = po.get('consignee_gst', '36AAQCS3643G1Z2')

    # ── header cells ──────────────────────────────────────────────
    _w(ws, 3,  5, f'Tax invoice No: {inv_no}')
    _w(ws, 4,  5, f'Date of issue of Invoice :{inv_date}')
    _w(ws, 5,  5, f'P.O.NO.: {po["po_number"]}')
    _w(ws, 7,  5, f'Total No of Cartons:- {total_cartons}')
    _w(ws, 8,  5, f'Total No of pcs. {total_pcs}')

    # ── delivery address (rows 10–15, col A) ──────────────────────
    _w(ws, 10, 1, consignee)
    for i, line in enumerate(addr[:4], 11):
        _w(ws, i, 1, line)
    _w(ws, 15, 1, f'GSTIN :{cgst}')

    # ── carton rows ───────────────────────────────────────────────
    FIRST = 17
    TMPL = TEMPLATE_PKG_CARTON_ROWS

    carton_styles = _save_row_styles(ws, FIRST, 10)

    ws.delete_rows(FIRST, TMPL)
    ws.insert_rows(FIRST, len(cartons))

    for i, c in enumerate(cartons):
        row = FIRST + i
        _apply_row_styles(ws, row, carton_styles)
        ws.row_dimensions[row].height = 29.25
        _w(ws, row, 1, '')
        _w(ws, row, 2, c['no'])
        _w(ws, row, 3, 'OF')
        _w(ws, row, 4, total_cartons)
        _w(ws, row, 5, c['article_size'])
        _w(ws, row, 6, c['description'])
        _w(ws, row, 7, c['color'])
        _w(ws, row, 8, c['qty'])
        _w(ws, row, 9, 1)
        _w(ws, row, 10, c['qty'])

    # ── TOTAL row (position shifts with carton count) ─────────────
    TOTAL = FIRST + len(cartons)
    _w(ws, TOTAL, 8, total_pcs)
    _w(ws, TOTAL, 10, total_pcs)


# ── public entry point ─────────────────────────────────────────────────────

def generate_excel(po_data, invoice_number, carton_capacity, invoice_date):
    wb = load_workbook(TEMPLATE_PATH)
    ws_inv = wb['INV']
    ws_pkg = wb['PKG']

    # Remove the stray medium-right border on col O (15) that comes from the template
    _clear_col_borders(ws_inv, 15)

    cartons = _calculate_cartons(po_data['items'], carton_capacity)
    total_cartons = len(cartons)

    _fill_inv(ws_inv, po_data, invoice_number, invoice_date, total_cartons)
    _fill_pkg(ws_pkg, po_data, invoice_number, invoice_date, cartons, total_cartons)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
